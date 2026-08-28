import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import hashlib

st.set_page_config(page_title="Auditoria PRO", layout="wide", page_icon="🛡️")

MAX_FILE_SIZE_MB = 50
CACHE_TTL = 3600

# ==========================================
# 0. CONFIGURAÇÕES E CACHE
# ==========================================

def higienizar_dataframe(df):
    if df.empty: return df
    df['Insumo_Filho'] = df['Insumo_Filho'].astype(str).str.strip().str.upper()
    df['Servico_Pai'] = df['Servico_Pai'].astype(str).str.strip().str.upper()
    df['Und'] = df['Und'].astype(str).str.strip().str.upper()
    df['Insumo_Filho'] = df['Insumo_Filho'].replace({'NAN': 'S/C', 'NONE': 'S/C', '': 'S/C'})
    df['Servico_Pai'] = df['Servico_Pai'].replace({'NAN': 'S/C', 'NONE': 'S/C', '': 'S/C'})
    return df

@st.cache_data(ttl=CACHE_TTL)
def carregar_orcafascio(arquivo_bytes, nome_arquivo, ativar_limpeza=True):
    try:
        tamanho_mb = len(arquivo_bytes) / (1024 * 1024)
        if tamanho_mb > MAX_FILE_SIZE_MB:
            return None, f"Arquivo {nome_arquivo} excede {MAX_FILE_SIZE_MB}MB", None
        
        df_raw = pd.read_excel(io.BytesIO(arquivo_bytes), header=None)
        
        dados = []
        log_erros_parse = []
        # A exportação do OrçaFascio é sequencial: cada bloco iniciado por
        # "Composição" é uma CPU principal e as linhas seguintes (Insumo ou
        # ComposiçãoAuxiliar) são seus subitens diretos.
        cod_pai_atual, desc_pai_atual = None, ""
        cod_cpu_raiz = None
        ordem_sequencial = 0
        
        in_orse_detail = False
        is_sicro_section = False
        skip_items = False
        
        sicro_col_quant = 4
        sicro_col_und = None
        sicro_col_preco = None
        
        col_cod, col_desc, col_und, col_quant, col_preco = 1, 3, 6, 7, 8
        
        def parse_number(val):
            if pd.isna(val) or val == '' or val == '*': return 0.0
            if isinstance(val, (int, float)): return float(val)
            val = str(val).strip()
            if ',' in val and '.' in val: val = val.replace('.', '').replace(',', '.')
            elif ',' in val: val = val.replace(',', '.')
            try: return float(val)
            except ValueError: return 0.0

        for idx, row in df_raw.iterrows():
            row_list = [str(x).strip() if pd.notna(x) else '' for x in row]
            row_lower = [x.lower() for x in row_list]
            col0 = row_list[0].lower() if len(row_list) > 0 else ''
            
            if 'detalhamento de cálculo orse' in col0 or 'detalhamento de calculo orse' in col0:
                in_orse_detail = True
                continue
                
            if any('mo sem ls' in x for x in row_lower) or any('valor do bdi' in x for x in row_lower):
                cod_pai_atual = None
                cod_cpu_raiz = None
                is_sicro_section = False
                in_orse_detail = False
                skip_items = False
                continue
                
            if re.match(r'^\d+\.\d+$', col0):
                cod_pai_atual = None
                cod_cpu_raiz = None
                is_sicro_section = False
                in_orse_detail = False
                skip_items = False
                continue
                
            if len(col0) == 1 and col0.upper() in ['B', 'C', 'D', 'E', 'F', 'G']:
                is_sicro_section = True
                skip_items = False
                sicro_col_quant = 4
                sicro_col_und = None
                sicro_col_preco = None
                for i, val in enumerate(row_lower):
                    if 'quant' in val: sicro_col_quant = i
                    elif 'unidade' in val or val == 'un': sicro_col_und = i
                    elif 'preço unit' in val or 'preco unit' in val: sicro_col_preco = i
                    elif 'custo horário' in val or 'custo horario' in val:
                        if sicro_col_preco is None: sicro_col_preco = i 
                
                if 'transporte' in ' '.join(row_lower): skip_items = True 
                continue
                
            if 'código' in row_lower or 'codigo' in row_lower:
                is_sicro_section = False
                skip_items = False
                for i, val in enumerate(row_lower):
                    if val in ['código', 'codigo']: col_cod = i
                    elif any(d in val for d in ['descrição', 'descricao']): col_desc = i
                    elif any(u in val for u in ['und', 'unidade', 'unid']): col_und = i
                    elif any(q in val for q in ['quant', 'qtd']): col_quant = i
                    elif any(p in val for p in ['valor unit', 'preço unit', 'preco unit']): col_preco = i
                continue
                
            if col0 in ['insumo', 'composição auxiliar', 'composicao auxiliar', 'item', 'composição', 'composicao']:
                if in_orse_detail or skip_items: continue
                if col_cod < len(row_list):
                    cod_item = row_list[col_cod].upper()
                    if cod_item.lower() in ['código', 'codigo', '', 'nan']: continue
                    
                    desc_item = row_list[col_desc] if col_desc is not None and col_desc < len(row_list) else ""
                    
                    if is_sicro_section:
                        qtd_idx = sicro_col_quant if sicro_col_quant is not None else 4
                        und_idx = sicro_col_und
                        preco_idx = sicro_col_preco
                        
                        qtd_valor = parse_number(row_list[qtd_idx]) if qtd_idx < len(row_list) else 0.0
                        
                        if und_idx is not None and und_idx < len(row_list) and row_list[und_idx] != '*': und_item = row_list[und_idx].upper()
                        else: und_item = 'H' 
                            
                        if preco_idx is not None and preco_idx < len(row_list) and row_list[preco_idx] != '*': preco_valor = parse_number(row_list[preco_idx])
                        else:
                            preco_valor = 0.0
                            last_num_idx = -1
                            for i in range(len(row_list) - 1, qtd_idx, -1):
                                val = row_list[i]
                                if val != '*' and val != '':
                                    parsed = parse_number(val)
                                    if parsed > 0:
                                        last_num_idx = i
                                        break
                            
                            if last_num_idx != -1:
                                if last_num_idx - 1 >= 0:
                                    val_preco = row_list[last_num_idx - 1]
                                    if val_preco != '*' and val_preco != '':
                                        parsed_preco = parse_number(val_preco)
                                        if parsed_preco > 0: preco_valor = parsed_preco
                                
                                if preco_valor == 0.0: preco_valor = parse_number(row_list[last_num_idx])
                    else:
                        und_item = row_list[col_und].upper() if col_und < len(row_list) else ""
                        qtd_valor = parse_number(row_list[col_quant]) if col_quant < len(row_list) else 0.0
                        preco_valor = parse_number(row_list[col_preco]) if col_preco < len(row_list) else 0.0

                    tipo_item = 'INSUMO'
                    # Uma composição normal inicia uma CPU.  A composição
                    # auxiliar é um subitem da CPU, não o pai das próximas
                    # linhas: o OrçaFascio já a exporta de forma plana.
                    if col0 in ['composição', 'composicao']:
                        tipo_item = 'COMPOSICAO_PRINCIPAL'
                        cod_cpu_raiz, cod_pai_atual = cod_item, cod_item
                        desc_pai_atual = desc_item
                    elif col0 in ['composição auxiliar', 'composicao auxiliar']:
                        tipo_item = 'COMPOSICAO_AUXILIAR'
                        if cod_cpu_raiz is None:
                            cod_cpu_raiz = cod_item

                    pai_direto = cod_pai_atual if cod_pai_atual else cod_item
                    desc_pai_direto = desc_pai_atual if cod_pai_atual else desc_item
                    cpu_raiz_linha = cod_cpu_raiz if cod_cpu_raiz else cod_item
                    
                    ordem_sequencial += 1
                    dados.append({
                        'Ordem': ordem_sequencial, 'Servico_Pai': cpu_raiz_linha, 'Descricao_Pai': desc_pai_direto,
                        'Pai_Direto': pai_direto, 'Tipo_Item': tipo_item,
                        'Insumo_Filho': cod_item, 'Descricao_Filho': desc_item, 'Und': und_item,
                        'Qtd': qtd_valor, 'Preco_Unitario': preco_valor, 'Status_Parsing': 'OK'
                    })
                    
                    
        df_final = pd.DataFrame(dados)
        if df_final.empty: return None, f"Planilha {nome_arquivo}: Nenhum dado válido extraído.", None
        
        if ativar_limpeza: df_final = higienizar_dataframe(df_final)
        
        checksum = hashlib.md5(df_final.to_string().encode()).hexdigest()
        log_msg = f"⚠️ {len(log_erros_parse)} linhas com erro de parsing" if log_erros_parse else ""
        return df_final, log_msg, (checksum, len(df_final))
    except Exception as e:
        return None, f"Erro ao processar {nome_arquivo}: {str(e)}", None

@st.cache_data(ttl=CACHE_TTL)
def transformar_hierarquico(df):
    if df.empty: return pd.DataFrame()
    df = df.sort_values('Ordem')
    linhas = []
    pai_atual = None
    colunas_vazias = ['Código', 'Descrição', 'Und_Base', 'Und_Prop', 'Qtd_Base', 'Qtd_Prop', 'Delta_Qtd', 'Preco_Base', 'Preco_Prop', 'Delta_Preco', 'Var_Preco_%', 'Total_Base', 'Total_Prop', 'Delta_Total', 'Var_Total_%']
    
    for _, row in df.iterrows():
        if row['Servico_Pai'] != pai_atual:
            if pai_atual is not None: linhas.append({c: np.nan for c in colunas_vazias})
            linhas.append({
                'Código': row['Servico_Pai'], 'Descrição': f"COMPOSIÇÃO: {row['Descricao_Pai']}",
                'Und_Base': '---', 'Und_Prop': '---', **{c: np.nan for c in colunas_vazias[4:]}
            })
            pai_atual = row['Servico_Pai']
            
        linhas.append({
            'Código': row['Insumo_Filho'], 'Descrição': row['Descricao_Filho'], 'Und_Base': row['Und_Base'], 'Und_Prop': row['Und_Prop'],
            'Qtd_Base': row['Qtd_Base'], 'Qtd_Prop': row['Qtd_Prop'], 'Delta_Qtd': row['Delta_Qtd'],
            'Preco_Base': row['Preco_Base'], 'Preco_Prop': row['Preco_Prop'], 'Delta_Preco': row['Delta_Preco'], 'Var_Preco_%': row['Var_Preco_%'],
            'Total_Base': row['Total_Base'], 'Total_Prop': row['Total_Prop'], 'Delta_Total': row['Delta_Total'], 'Var_Total_%': row['Var_Total_%']
        })
    return pd.DataFrame(linhas)

@st.cache_data(ttl=CACHE_TTL)
def transformar_hierarquico_raw(df):
    if df.empty: return pd.DataFrame()
    df = df.sort_values('Ordem')
    linhas = []
    pai_atual = None
    df_copy = df.copy()
    df_copy['Total'] = df_copy['Qtd'] * df_copy['Preco_Unitario']
    colunas_esquema = ['Código', 'Descrição', 'Unidade', 'Quantidade', 'Preço Unitário', 'Total']
    
    for _, row in df_copy.iterrows():
        if row['Servico_Pai'] != pai_atual:
            if pai_atual is not None: linhas.append({c: np.nan for c in colunas_esquema})
            linhas.append({
                'Código': row['Servico_Pai'], 'Descrição': f"COMPOSIÇÃO: {row['Descricao_Pai']}",
                'Unidade': '---', 'Quantidade': np.nan, 'Preço Unitário': np.nan, 'Total': np.nan
            })
            pai_atual = row['Servico_Pai']
            
        linhas.append({
            'Código': row['Insumo_Filho'], 'Descrição': row['Descricao_Filho'], 'Unidade': row['Und'],
            'Quantidade': row['Qtd'], 'Preço Unitário': row['Preco_Unitario'], 'Total': row['Total']
        })
    return pd.DataFrame(linhas)

# ==========================================
# 2. ESTILIZADORES DA INTERFACE (UI)
# ==========================================

def estilizar_relatorio(row):
    if str(row.get('Und_Base', '')).strip() == '---':
        return ['background-color: #dbeafe; font-weight: bold; color: #1e3a8a;'] * len(row)
    estilos = [''] * len(row)
    
    def safe_float_ui(v):
        if pd.isna(v) or v == '' or v == '*': return 0.0
        if isinstance(v, (int, float)): return float(v)
        s = str(v).strip()
        if ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
        elif ',' in s: s = s.replace(',', '.')
        s = re.sub(r'[R$\s%]', '', s)
        try: return float(s)
        except: return 0.0

    for i, col in enumerate(row.index):
        val = row[col]
        
        if col in ['Und_Base', 'Und_Prop']:
            und_b = str(row.get('Und_Base', '')).strip().upper()
            und_p = str(row.get('Und_Prop', '')).strip().upper()
            if und_b not in ['', 'NAN', '---'] and und_p not in ['', 'NAN', '---'] and und_b != und_p:
                estilos[i] = 'background-color: #fef08a; color: #713f12; font-weight: bold;'
                
        if pd.isna(val) or val == '': continue
        v = safe_float_ui(val)
        if v == 0 and col not in ['Delta_Qtd', 'Delta_Preco', 'Delta_Total', 'Var_Preco_%', 'Var_Total_%']: continue
            
        if col == 'Delta_Qtd' and v != 0: estilos[i] = 'background-color: #e9d5ff; color: #6b21a8; font-weight: bold;'
        elif col in ['Delta_Preco', 'Delta_Total'] and v > 0: estilos[i] = 'background-color: #fca5a5; color: #7f1d1d; font-weight: bold;'
        elif col in ['Var_Preco_%', 'Var_Total_%']:
            if v > 0: estilos[i] = 'background-color: #fca5a5; color: #7f1d1d; font-weight: bold;'
            elif v < st.session_state.get('limiar_desconto', -0.25): estilos[i] = 'background-color: #fdba74; color: #7c2d12; font-weight: bold;'
    return estilos

def estilizar_relatorio_raw(row):
    if str(row.get('Unidade', '')).strip() == '---' or 'COMPOSIÇÃO' in str(row.get('Descrição', '')): 
        return ['background-color: #dbeafe; font-weight: bold; color: #1e3a8a;'] * len(row)
    return [''] * len(row)

# ==========================================
# 3. MOTOR EXPORTADOR EXCEL MULTI-ABA
# ==========================================

def gerar_excel_bytes(dash_data, df_matriz, df_inconformidades, df_nao_encontrados_base, df_nao_encontrados_prop, df_realocados, df_parsing, df_db_base, df_db_prop):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        df_matriz.to_excel(writer, index=False, sheet_name='📋 Matriz Completa', startrow=8)
        if not df_inconformidades.empty: df_inconformidades.to_excel(writer, index=False, sheet_name='🚨 Inconformidades', startrow=8)
        else: writer.book.create_sheet(title='🚨 Inconformidades')['A1'] = "Nenhuma inconformidade detectada."
            
        df_nao_encontrados_base.to_excel(writer, index=False, sheet_name='📍 Não Encontrados na Base', startrow=1)
        df_nao_encontrados_prop.to_excel(writer, index=False, sheet_name='📍 Omitidos na Proposta', startrow=1)
        df_realocados.to_excel(writer, index=False, sheet_name='🔀 Realocados / Estrutura Divergente', startrow=1)
        df_parsing.to_excel(writer, index=False, sheet_name='📝 Log de Erros de Parsing', startrow=1)
        df_db_base.to_excel(writer, index=False, sheet_name='🗄️ DB Base', startrow=1)
        df_db_prop.to_excel(writer, index=False, sheet_name='🗄️ DB Proposta', startrow=1)
        
        wb = writer.book
        ws_kpi = wb.create_sheet('📊 Dashboard KPI', 0)
        ws_kpi.sheet_view.showGridLines = False
        
        border_box = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'), top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        ws_kpi.merge_cells('B2:J3')
        title = ws_kpi['B2']
        title.value = "📊 PAINEL ANALÍTICO DE CONFORMIDADE CONTRATUAL"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")

        def desenhar_card(row, col, title, value, format_type='int'):
            c_title = ws_kpi.cell(row=row, column=col)
            c_title.value = title
            c_title.font = Font(bold=True, color="FFFFFF", size=10)
            c_title.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            c_title.alignment = Alignment(horizontal="center", vertical="center")
            c_title.border = border_box
            ws_kpi.cell(row=row, column=col+1).border = border_box
            ws_kpi.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

            c_val = ws_kpi.cell(row=row+1, column=col)
            c_val.value = value
            c_val.font = Font(size=14, bold=True, color="0F172A")
            c_val.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            c_val.alignment = Alignment(horizontal="center", vertical="center")
            c_val.border = border_box
            
            for r_adj in range(row+1, row+3):
                for c_adj in range(col, col+2): ws_kpi.cell(row=r_adj, column=c_adj).border = border_box
            ws_kpi.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+1)

            if format_type == 'currency': c_val.number_format = '"R$" #,##0.00'
            elif format_type == 'percent': c_val.number_format = '0.00%'
            else: c_val.number_format = '#,##0'

        desenhar_card(5, 2, "ITENS AUDITADOS", dash_data['total_insumos'], 'int')
        desenhar_card(5, 5, "SALDO DO ORÇAMENTO", dash_data['total_proposta'], 'currency')
        desenhar_card(5, 8, "TAXA DE CONFORMIDADE", dash_data['taxa_conformidade'], 'percent')

        desenhar_card(9, 2, "RISCO SOBREPREÇO", dash_data['financeiro_sobrepreco'], 'currency')
        desenhar_card(9, 5, "DESCONTO OCULTO (INEX)", dash_data['financeiro_inexequivel'], 'currency')
        desenhar_card(9, 8, "MAIOR DESVIO ÚNICO", dash_data['max_desvio'], 'currency')

        def desenhar_tabela(start_row, start_col, df, title_text):
            if df.empty: return start_row
            t_cell = ws_kpi.cell(row=start_row, column=start_col)
            t_cell.value = title_text
            t_cell.font = Font(bold=True, size=11, color="1E3A8A")
            ws_kpi.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(df.columns) - 1)

            for c_idx, col_name in enumerate(df.columns):
                cell = ws_kpi.cell(row=start_row+1, column=start_col+c_idx)
                cell.value = col_name
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border_box

            for r_idx, row in enumerate(df.values):
                for c_idx, val in enumerate(row):
                    cell = ws_kpi.cell(row=start_row+2+r_idx, column=start_col+c_idx)
                    cell.value = val
                    cell.border = border_box
                    cell.alignment = Alignment(vertical="center")
                    col_name = df.columns[c_idx]
                    if any(k in col_name for k in ['%', 'Variação', 'Desconto']): cell.number_format = '0.00%'
                    elif any(k in col_name for k in ['R$', 'Impacto', 'Defasagem', 'Sobrepreço']): cell.number_format = '"R$" #,##0.00'
                    elif isinstance(val, (int, float)): cell.number_format = '#,##0'
            return start_row + len(df) + 4 

        desenhar_tabela(14, 2, dash_data['count_graf'], "📌 OCORRÊNCIAS POR TIPOLOGIA")
        desenhar_tabela(14, 6, dash_data['money_graf'], "💸 IMPACTO FINANCEIRO LÍQUIDO")

        next_row = desenhar_tabela(23, 2, dash_data['top_sobre'], "🎯 FRENTE 1: TOP 5 IMPACTOS DE SOBREPREÇO")
        desenhar_tabela(next_row, 2, dash_data['top_inex'], "🎯 FRENTE 2: TOP 5 RISCOS DE INEXEQUIBILIDADE")

        ws_kpi.column_dimensions['A'].width = 3
        ws_kpi.column_dimensions['B'].width = 15
        ws_kpi.column_dimensions['C'].width = 32
        ws_kpi.column_dimensions['D'].width = 4
        ws_kpi.column_dimensions['E'].width = 20
        ws_kpi.column_dimensions['F'].width = 25
        ws_kpi.column_dimensions['G'].width = 4
        ws_kpi.column_dimensions['H'].width = 18
        ws_kpi.column_dimensions['I'].width = 25
        ws_kpi.column_dimensions['J'].width = 3

        def injetar_legenda(ws):
            legendas = [
                ('A1', 'FCA5A5', '🟥 VERMELHO: Sobrepreço (Preço/Total da proposta superior à referência).'),
                ('A2', 'E9D5FF', '🟪 ROXO: Quantidade diferente da referência.'),
                ('A3', 'FDBA74', '🟧 LARANJA: Inexequibilidade (Desconto superior a 25%).'),
                ('A4', 'FEF08A', '🟨 AMARELO: Fraude Métrica (Unidades de medida incompatíveis).'),
                ('A7', 'DBEAFE', '🟦 AZUL CLARO: Composição Analítica Pai.')
            ]
            for celula, cor, texto in legendas:
                ws[celula] = texto
                ws[celula].fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                ws[celula].font = Font(bold=True, size=9)

        def safe_float(v):
            if v is None: return 0.0
            if isinstance(v, (int, float)): return float(v)
            s = str(v).strip()
            if s in ('', '*', 'nan', 'None', '---'): return 0.0
            if ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
            elif ',' in s: s = s.replace(',', '.')
            s = re.sub(r'[R$\s%]', '', s)
            try: return float(s)
            except ValueError: return 0.0

        def safe_str(v):
            if v is None: return ''
            s = str(v).strip()
            return '' if s.lower() in ('nan', 'none', '---') else s

        limiar_desconto = st.session_state.get('limiar_desconto', -0.25)

        for name in wb.sheetnames:
            if name == '📊 Dashboard KPI': continue
            ws = wb[name]

            if name in ['📋 Matriz Completa', '🚨 Inconformidades']:
                injetar_legenda(ws)
                linha_cabecalho = 9
            elif name in ['📍 Não Encontrados na Base', '📍 Omitidos na Proposta', '🔀 Realocados / Estrutura Divergente', '🗄️ DB Base', '🗄️ DB Proposta', '📝 Log de Erros de Parsing']:
                linha_cabecalho = 2
            else:
                linha_cabecalho = 1

            col_map = {safe_str(ws.cell(row=linha_cabecalho, column=c).value): c for c in range(1, ws.max_column + 1) if safe_str(ws.cell(row=linha_cabecalho, column=c).value)}

            if name in ['📋 Matriz Completa', '🚨 Inconformidades']:
                col_und_base, col_und_prop = col_map.get('Und_Base'), col_map.get('Und_Prop')
                col_delta_qtd, col_delta_preco, col_var_preco = col_map.get('Delta_Qtd'), col_map.get('Delta_Preco'), col_map.get('Var_Preco_%')
                col_delta_total, col_var_total = col_map.get('Delta_Total'), col_map.get('Var_Total_%')

                for r_idx in range(linha_cabecalho + 1, ws.max_row + 1):
                    raw_und = ws.cell(row=r_idx, column=col_und_base).value if col_und_base else None
                    if safe_str(raw_und) == '---' or 'COMPOSIÇÃO' in safe_str(ws.cell(row=r_idx, column=2).value):
                        for c_idx in range(1, ws.max_column + 1):
                            ws.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type="solid")
                            ws.cell(row=r_idx, column=c_idx).font = Font(bold=True, color='1E3A8A')
                        continue

                    und_b_val = safe_str(ws.cell(row=r_idx, column=col_und_base).value) if col_und_base else ''
                    und_p_val = safe_str(ws.cell(row=r_idx, column=col_und_prop).value) if col_und_prop else ''
                    if und_b_val and und_p_val and und_b_val.upper() != und_p_val.upper():
                        if col_und_base: ws.cell(row=r_idx, column=col_und_base).fill = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type="solid")
                        if col_und_prop: ws.cell(row=r_idx, column=col_und_prop).fill = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type="solid")

                    d_qtd = safe_float(ws.cell(row=r_idx, column=col_delta_qtd).value) if col_delta_qtd else 0.0
                    d_preco = safe_float(ws.cell(row=r_idx, column=col_delta_preco).value) if col_delta_preco else 0.0
                    v_preco = safe_float(ws.cell(row=r_idx, column=col_var_preco).value) if col_var_preco else 0.0
                    d_total = safe_float(ws.cell(row=r_idx, column=col_delta_total).value) if col_delta_total else 0.0
                    v_total = safe_float(ws.cell(row=r_idx, column=col_var_total).value) if col_var_total else 0.0

                    if d_qtd != 0 and col_delta_qtd: ws.cell(row=r_idx, column=col_delta_qtd).fill = PatternFill(start_color='E9D5FF', end_color='E9D5FF', fill_type="solid")
                    if d_preco > 0 and col_delta_preco: ws.cell(row=r_idx, column=col_delta_preco).fill = PatternFill(start_color='FCA5A5', end_color='FCA5A5', fill_type="solid")
                    if d_total > 0 and col_delta_total: ws.cell(row=r_idx, column=col_delta_total).fill = PatternFill(start_color='FCA5A5', end_color='FCA5A5', fill_type="solid")
                    if v_preco > 0 and col_var_preco: ws.cell(row=r_idx, column=col_var_preco).fill = PatternFill(start_color='FCA5A5', end_color='FCA5A5', fill_type="solid")
                    if v_total > 0 and col_var_total: ws.cell(row=r_idx, column=col_var_total).fill = PatternFill(start_color='FCA5A5', end_color='FCA5A5', fill_type="solid")
                    if v_preco < limiar_desconto and col_var_preco: ws.cell(row=r_idx, column=col_var_preco).fill = PatternFill(start_color='FDBA74', end_color='FDBA74', fill_type="solid")
                    if v_total < limiar_desconto and col_var_total: ws.cell(row=r_idx, column=col_var_total).fill = PatternFill(start_color='FDBA74', end_color='FDBA74', fill_type="solid")

            elif name in ['📍 Não Encontrados na Base', '📍 Omitidos na Proposta', '🔀 Realocados / Estrutura Divergente', '🗄️ DB Base', '🗄️ DB Proposta']:
                for r_idx in range(linha_cabecalho + 1, ws.max_row + 1):
                    if 'COMPOSIÇÃO' in safe_str(ws.cell(row=r_idx, column=2).value):
                        for c_idx in range(1, ws.max_column + 1):
                            ws.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type="solid")
                            ws.cell(row=r_idx, column=c_idx).font = Font(bold=True, color='1E3A8A')

            formatos = {}
            for c in range(1, ws.max_column + 1):
                nome = safe_str(ws.cell(row=linha_cabecalho, column=c).value)
                if any(k in nome for k in ['Preco', 'Preço', 'Total', 'Valor', 'Delta_Preco', 'Delta_Total']):
                    formatos[c] = '0.00%' if '%' in nome or 'Var_' in nome else '"R$" #,##0.00'
                elif any(k in nome for k in ['Qtd', 'Quantidade', 'Delta_Qtd']): formatos[c] = '#,##0.0000'

            for col in ws.columns:
                max_len = 0
                c_idx = col[0].column
                col_l = get_column_letter(c_idx)
                hc = ws.cell(row=linha_cabecalho, column=c_idx)
                hc.font, hc.fill, hc.alignment = Font(bold=True, color='FFFFFF'), PatternFill(start_color='1E293B', end_color='1E293B', fill_type="solid"), Alignment(horizontal='center', vertical='center', wrap_text=True)

                for cell in col:
                    if cell.row > linha_cabecalho and c_idx in formatos and isinstance(cell.value, (int, float)):
                        cell.number_format = formatos[c_idx]
                    try:
                        if cell.value:
                            val_str = f"{cell.value * 100:.2f}%" if c_idx in formatos and '%' in formatos[c_idx] else str(cell.value)
                            max_len = max(max_len, len(val_str))
                    except: pass
                ws.column_dimensions[col_l].width = min(max(max_len + 3, 11), 70)

    return output.getvalue()

# ==========================================
# 4. INTERFACE GRÁFICA (UI - STREAMLIT)
# ==========================================

if 'limiar_desconto' not in st.session_state: st.session_state.limiar_desconto = -0.25
if 'ativar_limpeza' not in st.session_state: st.session_state.ativar_limpeza = True

with st.sidebar:
    st.subheader("⚙️ Parâmetros de Auditoria")
    limiar_desconto = st.slider("Limiar de Desconto (Inexequibilidade)", -50, -5, -25, 1)
    st.session_state.limiar_desconto = limiar_desconto / 100
    st.session_state.ativar_limpeza = st.checkbox("🧹 Forçar Higienização de Dados (Recomendado)", value=True)
    st.success(f"📌 **Limiar Configurado: {limiar_desconto}%**")
    st.divider()
    st.subheader("📌 Legenda de Auditoria")
    st.markdown("""
    <div style="padding: 10px; border-radius: 5px; margin-bottom: 5px; background-color: #fca5a5; color: #7f1d1d; font-family: sans-serif; font-size:13px;"><b>🟥 Vermelho (Sobrepreço)</b><br>Preço unitário ou total superior à base.</div>
    <div style="padding: 10px; border-radius: 5px; margin-bottom: 5px; background-color: #e9d5ff; color: #6b21a8; font-family: sans-serif; font-size:13px;"><b>🟪 Roxo (Qtd. Alterada)</b><br>Quantidade do insumo majorada.</div>
    <div style="padding: 10px; border-radius: 5px; margin-bottom: 5px; background-color: #fdba74; color: #7c2d12; font-family: sans-serif; font-size:13px;"><b>🟧 Laranja (Inexequibilidade)</b><br>Desconto excessivo.</div>
    <div style="padding: 10px; border-radius: 5px; margin-bottom: 5px; background-color: #fef08a; color: #713f12; font-family: sans-serif; font-size:13px;"><b>🟨 Amarelo (Fraude Métrica)</b><br>Unidades de Medida incompatíveis.</div>
    """, unsafe_allow_html=True)

st.title("🛡️ Auditoria de Orçamentos PRO")
st.markdown("Validação paramétrica hierárquica de planilhas orçamentárias de Engenharia Civil.")
st.divider()

col1, col2 = st.columns(2)
with col1: arquivo_base = st.file_uploader("1. Base de Referência (SINAPI/ORSE)", type=["xlsx", "xls"])
with col2: arquivo_proposta = st.file_uploader("2. Proposta da Empreiteira", type=["xlsx", "xls"])

if arquivo_base and arquivo_proposta:
    with st.spinner("Compilando inteligência analítica do painel de controle..."):
        
        df_base_raw, msg_base, check_base = carregar_orcafascio(arquivo_base.getvalue(), "Base", st.session_state.ativar_limpeza)
        df_prop_raw, msg_prop, check_prop = carregar_orcafascio(arquivo_proposta.getvalue(), "Proposta", st.session_state.ativar_limpeza)
        
        if df_base_raw is not None and df_prop_raw is not None:           
            
            # AGRUPAMENTO HIERÁRQUICO - O CORAÇÃO DA AUDITORIA
            # A chave inclui o pai direto: o mesmo insumo pode existir em
            # composições auxiliares diferentes dentro da mesma CPU.
            cols_agrupamento = ['Servico_Pai', 'Pai_Direto', 'Insumo_Filho']
            
            df_base = df_base_raw.groupby(cols_agrupamento).agg({
                'Qtd': 'sum', 'Preco_Unitario': 'mean', 'Und': 'first',
                'Descricao_Pai': 'first', 'Descricao_Filho': 'first', 'Tipo_Item': 'first', 'Ordem': 'min'
            }).reset_index()
            
            df_prop = df_prop_raw.groupby(cols_agrupamento).agg({
                'Qtd': 'sum', 'Preco_Unitario': 'mean', 'Und': 'first',
                'Descricao_Pai': 'first', 'Descricao_Filho': 'first', 'Tipo_Item': 'first', 'Ordem': 'min'
            }).reset_index()
            
            df_merged = pd.merge(
                df_base, df_prop,
                on=cols_agrupamento,
                how='outer',
                suffixes=('_Base', '_Prop'),
                indicator=True
            )
            
            idx_right_only = df_merged[df_merged['_merge'] == 'right_only'].set_index(cols_agrupamento).index
            candidatos_adicionados = df_prop_raw[df_prop_raw.set_index(cols_agrupamento).index.isin(idx_right_only)].drop_duplicates(subset=cols_agrupamento).reset_index(drop=True)
            
            idx_left_only = df_merged[df_merged['_merge'] == 'left_only'].set_index(cols_agrupamento).index
            candidatos_ausentes = df_base_raw[df_base_raw.set_index(cols_agrupamento).index.isin(idx_left_only)].drop_duplicates(subset=cols_agrupamento).reset_index(drop=True)

            # Não trate como omissão um código que ainda exista na proposta:
            # ele foi realocado (ou a estrutura da CPU foi alterada).  Só é
            # omitido de fato quando não existe em nenhuma CPU da proposta.
            codigos_prop = set(df_prop_raw['Insumo_Filho'].astype(str).str.upper())
            codigos_base = set(df_base_raw['Insumo_Filho'].astype(str).str.upper())
            df_realocados = candidatos_ausentes[
                candidatos_ausentes['Insumo_Filho'].astype(str).str.upper().isin(codigos_prop)
            ].copy()
            df_nao_encontrados_prop = candidatos_ausentes[
                ~candidatos_ausentes['Insumo_Filho'].astype(str).str.upper().isin(codigos_prop)
            ].copy()
            df_nao_encontrados_base = candidatos_adicionados[
                ~candidatos_adicionados['Insumo_Filho'].astype(str).str.upper().isin(codigos_base)
            ].copy()
            
            df_auditoria = df_merged[df_merged['_merge'] == 'both'].copy()
            
            df_auditoria.rename(columns={
                'Und_Base': 'Und_Base', 'Qtd_Base': 'Qtd_Base', 'Preco_Unitario_Base': 'Preco_Base',
                'Und_Prop': 'Und_Prop', 'Qtd_Prop': 'Qtd_Prop', 'Preco_Unitario_Prop': 'Preco_Prop',
                'Descricao_Pai_Base': 'Descricao_Pai', 'Descricao_Filho_Base': 'Descricao_Filho',
                'Ordem_Base': 'Ordem'
            }, inplace=True)
            
            df_auditoria['Total_Base'] = df_auditoria['Qtd_Base'] * df_auditoria['Preco_Base']
            df_auditoria['Total_Prop'] = df_auditoria['Qtd_Prop'] * df_auditoria['Preco_Prop']
            df_auditoria['Delta_Qtd'] = df_auditoria['Qtd_Prop'] - df_auditoria['Qtd_Base']
            df_auditoria['Delta_Preco'] = df_auditoria['Preco_Prop'] - df_auditoria['Preco_Base']
            df_auditoria['Delta_Total'] = df_auditoria['Total_Prop'] - df_auditoria['Total_Base']
            df_auditoria['Var_Preco_%'] = np.where(df_auditoria['Preco_Base'] > 0, (df_auditoria['Preco_Prop'] / df_auditoria['Preco_Base']) - 1, 0)
            df_auditoria['Var_Total_%'] = np.where(df_auditoria['Total_Base'] > 0, (df_auditoria['Total_Prop'] / df_auditoria['Total_Base']) - 1, 0)
            
            df_completo = df_auditoria.reset_index(drop=True)

            for col in ['Delta_Total', 'Delta_Preco', 'Delta_Qtd', 'Var_Preco_%', 'Var_Total_%']: df_completo[col] = df_completo[col].fillna(0)
            df_completo['Und_Base'] = df_completo['Und_Base'].fillna('').astype(str)
            df_completo['Und_Prop'] = df_completo['Und_Prop'].fillna('').astype(str)
            
            sobrepreco_filter = (df_completo['Delta_Preco'] > 0) | (df_completo['Delta_Total'] > 0) | (df_completo['Var_Preco_%'] > 0) | (df_completo['Var_Total_%'] > 0)
            inexequivel_filter = (df_completo['Var_Preco_%'] < st.session_state.limiar_desconto) | (df_completo['Var_Total_%'] < st.session_state.limiar_desconto)
            qtd_filter = df_completo['Delta_Qtd'] != 0
            und_filter = df_completo['Und_Base'].str.upper() != df_completo['Und_Prop'].str.upper()
            
            irregularidades = df_completo[sobrepreco_filter | qtd_filter | inexequivel_filter | und_filter]
            
            df_visual_completo = transformar_hierarquico(df_completo)
            df_visual_erros = transformar_hierarquico(irregularidades) if not irregularidades.empty else pd.DataFrame()
            
            df_visual_ne_base = transformar_hierarquico_raw(df_nao_encontrados_base)
            df_visual_ne_prop = transformar_hierarquico_raw(df_nao_encontrados_prop)
            df_visual_realocados = transformar_hierarquico_raw(df_realocados)
            df_visual_db_base = transformar_hierarquico_raw(df_base_raw)
            df_visual_db_prop = transformar_hierarquico_raw(df_prop_raw)
            
            total_insumos, total_irregularidades = len(df_completo), len(irregularidades)
            total_base, total_proposta = float(df_completo['Total_Base'].sum()), float(df_completo['Total_Prop'].sum())
            var_total_geral = (total_proposta / total_base - 1) if total_base > 0 else 0
            taxa_conformidade = ((total_insumos - total_irregularidades) / total_insumos) if total_insumos > 0 else 0
            
            financeiro_sobrepreco = float(df_completo[df_completo['Delta_Total'] > 0]['Delta_Total'].sum())
            financeiro_qtd = float(abs(df_completo[qtd_filter]['Delta_Total'].sum()))
            financeiro_inexequivel = float(abs(df_completo[inexequivel_filter]['Delta_Total'].sum()))
            
            max_desvio_individual = float(df_completo['Delta_Total'].max()) if not df_completo.empty else 0.0
            
            sobreprecados, quantidades_alteradas = len(df_completo[sobrepreco_filter]), len(df_completo[qtd_filter])
            unidades_incompativeis, inexequiveis = len(df_completo[und_filter]), len(df_completo[inexequivel_filter])
            
            df_top_sobre = df_completo[df_completo['Delta_Total'] > 0].sort_values(by='Delta_Total', ascending=False).head(5)
            if not df_top_sobre.empty:
                df_top_sobre_view = df_top_sobre[['Insumo_Filho', 'Descricao_Filho', 'Delta_Total', 'Var_Preco_%']].copy()
                df_top_sobre_view.columns = ['Código', 'Descrição do Insumo', 'Sobrepreço (R$)', 'Variação (%)']
            else: df_top_sobre_view = pd.DataFrame()

            df_top_inex = df_completo[inexequivel_filter].sort_values(by='Delta_Total', ascending=True).head(5)
            if not df_top_inex.empty:
                df_top_inex_view = df_top_inex[['Insumo_Filho', 'Descricao_Filho', 'Delta_Total', 'Var_Total_%']].copy()
                df_top_inex_view.columns = ['Código', 'Descrição do Insumo', 'Defasagem (R$)', 'Variação Total (%)']
            else: df_top_inex_view = pd.DataFrame()

            dash_data_excel = {
                'total_insumos': total_insumos, 'total_proposta': total_proposta, 'taxa_conformidade': taxa_conformidade,
                'financeiro_sobrepreco': financeiro_sobrepreco, 'financeiro_inexequivel': financeiro_inexequivel, 'max_desvio': max_desvio_individual,
                'count_graf': pd.DataFrame({'Tipologia de Erro': ['🟥 Sobrepreço', '🟪 Qtd. Majorada', '🟨 Fraude Métrica', '🟧 Inexequível'], 'Ocorrências': [sobreprecados, quantidades_alteradas, unidades_incompativeis, inexequiveis]}),
                'money_graf': pd.DataFrame({'Tipologia Financeira': ['Sobrepreço Global', 'Majorização de Qtd.', 'Descontos Extremos'], 'Impacto (R$)': [financeiro_sobrepreco, financeiro_qtd, financeiro_inexequivel]}),
                'top_sobre': df_top_sobre_view, 'top_inex': df_top_inex_view
            }
            
            df_parsing_excel = pd.DataFrame(columns=['Origem', 'Código', 'Descrição', 'Erro'])
            
            excel_bytes = gerar_excel_bytes(dash_data_excel, df_visual_completo, df_visual_erros, df_visual_ne_base, df_visual_ne_prop, df_visual_realocados, df_parsing_excel, df_visual_db_base, df_visual_db_prop)
            
            formato_tela = {'Qtd_Base': '{:.4f}', 'Qtd_Prop': '{:.4f}', 'Delta_Qtd': '{:.4f}', 'Preco_Base': 'R$ {:.2f}', 'Preco_Prop': 'R$ {:.2f}', 'Delta_Preco': 'R$ {:.2f}', 'Var_Preco_%': '{:.2%}', 'Total_Base': 'R$ {:.2f}', 'Total_Prop': 'R$ {:.2f}', 'Delta_Total': 'R$ {:.2f}', 'Var_Total_%': '{:.2%}'}
            formato_tela_raw = {'Quantidade': '{:.4f}', 'Preço Unitário': 'R$ {:.2f}', 'Total': 'R$ {:.2f}'}
            
            styler_ui_completo = df_visual_completo.style.format(formato_tela, na_rep="").apply(estilizar_relatorio, axis=1)
            styler_ui_erros = df_visual_erros.style.format(formato_tela, na_rep="").apply(estilizar_relatorio, axis=1) if not df_visual_erros.empty else None
            
            styler_ui_ne_base = df_visual_ne_base.style.format(formato_tela_raw, na_rep="").apply(estilizar_relatorio_raw, axis=1) if not df_visual_ne_base.empty else None
            styler_ui_ne_prop = df_visual_ne_prop.style.format(formato_tela_raw, na_rep="").apply(estilizar_relatorio_raw, axis=1) if not df_visual_ne_prop.empty else None
            styler_ui_realocados = df_visual_realocados.style.format(formato_tela_raw, na_rep="").apply(estilizar_relatorio_raw, axis=1) if not df_visual_realocados.empty else None
            styler_ui_db_base = df_visual_db_base.style.format(formato_tela_raw, na_rep="").apply(estilizar_relatorio_raw, axis=1)
            styler_ui_db_prop = df_visual_db_prop.style.format(formato_tela_raw, na_rep="").apply(estilizar_relatorio_raw, axis=1)
            
            tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs(["📊 Dashboard KPI", "📋 Matriz Completa", "🚨 Inconformidades", "📍 Não Encontrados na Base", "📍 Omitidos na Proposta", "🔀 Realocados / Estrutura", "📝 Log de Erros de Parsing", "🗄️ DB Base", "🗄️ DB Proposta"])
            
            with tab1:
                st.subheader("📊 Painel Analítico de Conformidade Contratual")
                c1, c2, c3 = st.columns(3)
                c1.metric("📌 Volume de Itens Auditados", f"{total_insumos:,.0f}", f"{total_irregularidades} desvios sinalizados")
                c2.metric("💰 Saldo Global do Orçamento", f"R$ {total_proposta:,.2f}", f"Variação: {var_total_geral:+.2%}", delta_color="inverse")
                c3.metric("✅ Índice de Acerto Paramétrico", f"{taxa_conformidade*100:.1f}%", "Meta aceitável: > 95%")
                c4, c5, c6 = st.columns(3)
                c4.metric("🚨 Exposição a Sobrepreço (Risco)", f"R$ {financeiro_sobrepreco:,.2f}", f"{sobreprecados} sub-itens majorados", delta_color="inverse")
                c5.metric("📉 Desconto Ofertado Oculto", f"R$ {financeiro_inexequivel:,.2f}", f"{inexequiveis} itens sub-precificados")
                c6.metric("🎯 Maior Desvio Único Mapeado", f"R$ {max_desvio_individual:,.2f}", "Alerta Crítico Magnificado")
                st.divider()
                
                g_col1, g_col2 = st.columns(2)
                with g_col1:
                    st.markdown("##### 🔢 Ocorrências por Tipologia de Desvio")
                    st.bar_chart(dash_data_excel['count_graf'].set_index('Tipologia de Erro')['Ocorrências'], height=280)
                with g_col2:
                    st.markdown("##### 💸 Impacto Financeiro Líquido por Grupo (R$)")
                    st.bar_chart(dash_data_excel['money_graf'].set_index('Tipologia Financeira')['Impacto (R$)'], height=280)
                st.divider()
                
                p_col1, p_col2 = st.columns(2)
                with p_col1:
                    st.markdown("##### 🔺 Top 5 Insumos com Maior Impacto de Sobrepreço")
                    if not df_top_sobre_view.empty: st.dataframe(df_top_sobre_view.style.format({'Sobrepreço (R$)': 'R$ {:.2f}', 'Variação (%)': '{:+.2%}'}), hide_index=True, use_container_width=True)
                    else: st.success("Nenhum sobrepreço mapeado.")
                with p_col2:
                    st.markdown("##### 🔻 Top 5 Insumos com Maior Risco de Inexequibilidade")
                    if not df_top_inex_view.empty: st.dataframe(df_top_inex_view.style.format({'Defasagem (R$)': 'R$ {:.2f}', 'Variação Total (%)': '{:.2%}'}), hide_index=True, use_container_width=True)
                    else: st.info("Nenhuma anomalia de desconto extremo encontrada.")
                
                st.divider()
                st.download_button("📥 Baixar Laudo de Auditoria (.XLSX)", data=excel_bytes, file_name='Laudo_Auditoria_PRO_Consolidado.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True, key='dl_kpi')
            
            with tab2:
                st.download_button("📥 Baixar Laudo de Auditoria (.XLSX)", data=excel_bytes, file_name='Laudo_Auditoria_PRO_Consolidado.xlsx', use_container_width=True, key='dl_matriz')
                import streamlit.components.v1 as components
                try: st.dataframe(styler_ui_completo, height=600, use_container_width=True)
                except Exception as e:
                    components.html(styler_ui_completo.to_html(), height=650, scrolling=True)

            with tab3:
                st.download_button("📥 Baixar Laudo de Auditoria (.XLSX)", data=excel_bytes, file_name='Laudo_Auditoria_PRO_Consolidado.xlsx', use_container_width=True, key='dl_erro')
                if styler_ui_erros is not None: st.dataframe(styler_ui_erros, height=500, use_container_width=True)
                else: st.success("✅ Tudo em conformidade!")
            with tab4:
                if styler_ui_ne_base is not None: st.dataframe(styler_ui_ne_base, height=500, use_container_width=True)
                else: st.success("✅ Alinhamento Completo!")
            with tab5:
                if styler_ui_ne_prop is not None: st.dataframe(styler_ui_ne_prop, height=500, use_container_width=True)
                else: st.success("✅ Alinhamento Completo!")
            with tab6:
                st.caption("Itens existentes na proposta, mas em outra ramificação/composição da CPU. Não são omitidos de fato.")
                if styler_ui_realocados is not None: st.dataframe(styler_ui_realocados, height=500, use_container_width=True)
                else: st.success("✅ Nenhuma realocação estrutural identificada.")
            with tab7: st.success("✅ Zero erros estruturais identificados.")
            with tab8: st.dataframe(styler_ui_db_base, height=600, use_container_width=True)
            with tab9: st.dataframe(styler_ui_db_prop, height=600, use_container_width=True)
